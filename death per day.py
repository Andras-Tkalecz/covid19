from openpyxl import load_workbook
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt

def loadWorkbook(wbName):
    # load the workbook from file and get the first sheet
    wb = load_workbook(wbName)
    ws = wb.active
    
    df = pd.DataFrame(ws.values)
    df.fillna(0, inplace=True)
    df.columns = df.iloc[0,:]
    df = df.drop([0])
    df.set_index(df.columns[0], inplace=True)
    df.columns.name = ''
    
    return df

def rollingPlot(df, window):
    #create rolling average
    df_ra = df.rolling(window, center=True, min_periods=1).mean().round()
    
    #calculate week-wise differences
    df_rad = df_ra.copy()
    for i, col in enumerate(df_rad.columns[1:]):
        df_rad.iloc[:,i+1] = df_ra.iloc[:,i+1]-df_ra.iloc[:,i]
    
    #filter negative values
    df_rad[df_rad < 0] = 0
    
    #stacked-area plot
    matplotlib.style.use('seaborn-pastel')
    labelz = ["%s,  \u2020 %d" % (dat.strftime("%b %d"),df_rad[dat].sum()) for dat in df_rad.columns]
    axes1 = df_rad.plot.area(stacked=True, title='Deaths per day')
    h1, l1 = axes1.get_legend_handles_labels()
    axes1.legend(h1,labelz)
    plt.tight_layout()
    
    #separate week plots
    axes2 = df_rad.plot.area(stacked=False, title='Deaths per day', subplots=True)
    for i,axe in enumerate(axes2):
        h,l = axe.get_legend_handles_labels()
        axe.legend(h,[labelz[i]])
    
    plt.tight_layout()


def split_covid_exDeath(df):
    exDeathz = df.iloc[:,1:3]
    exDeath = df.iloc[:,2]-df.iloc[:,1]
    df.drop(df.columns[:3], axis=1, inplace=True)

    return df, exDeath, exDeathz


def exDeathOlayPlot(df, exDeath):
    fig, ax = plt.subplots()
    
    df_all = df.iloc[:-14,-1].rolling(7, center=True, min_periods=1).mean().round()
    df_all_axes = df_all.plot.area(stacked=True, title='Deaths per day', ax=ax)
    df_all_label = "covid-19,  \u2020 %d" % df_all.sum()
    
    exDed = exDeath[:-14].rolling(7, center=True, min_periods=1).mean().round()
    exDed_axes = exDed.plot(color='black', label='Excess Death', ax=ax)
    exDed_label = "excess death,  \u2020 %d" % exDed.sum()
    
    labelz = [exDed_label, df_all_label]
    h1, l1 = ax.get_legend_handles_labels()
    ax.legend(h1,labelz)
    
    ax.axis('auto')
    plt.tight_layout()


def exDeathsPlot(exDeathz):
    fig, ax = plt.subplots()
    
    exDeathz_ra = exDeathz[:-14].rolling(7, center=True, min_periods=1).mean().round()
    exDeathz_ra_axes = exDeathz_ra.plot.area(stacked=False, title='Excess Death', ax=ax)
    
    plt.tight_layout()


df = loadWorkbook('Death per day.xlsx')
df, exDeath, exDeathz = split_covid_exDeath(df)

rollingPlot(df,7)
exDeathOlayPlot(df, exDeath)
exDeathsPlot(exDeathz)

#every 2 weeks
df_e2v = df.copy()
df_e2v.drop(df_e2v.columns[1::2], axis=1, inplace=True)
rollingPlot(df_e2v,7)

#first and last week comparison
df_fnl = df.copy()
df_fnl.drop(df_fnl.columns[1:-1], axis=1, inplace=True)
rollingPlot(df_fnl,7)
rollingPlot(df[:-24],7)
