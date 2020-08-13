#USA covid and excess death comparison
from openpyxl import load_workbook
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt


def loadUsaWorkbook(wbName):
    # load the workbook from file and get the first sheet
    wb = load_workbook(wbName)
    ws = wb.active
    
    df = pd.DataFrame(ws.values)
    df.fillna(0, inplace=True)
    df.columns = df.iloc[0,:]
    df = df.drop([0])
    df.set_index(df.columns[1], inplace=True)
    df.columns.name = ''
    
    df.drop(df.columns[[0,1,2,3,4,12]], axis=1, inplace=True)
    
    return df


def plotUsaDeaths(df_USA):
    df_USA.drop(df_USA.columns[[2,3,4,5]], axis=1, inplace=True)
    df_USA_axes = df_USA[:-6].plot.area(stacked=False, title='Deaths')
    


df_USA = loadUsaWorkbook('USA covid excess death.xlsx')
df_USA.iloc[:,1] = df_USA.iloc[:,1] - df_USA.iloc[0,1]
plotUsaDeaths(df_USA)